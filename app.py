import os
import json
import io
from datetime import datetime
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, Response, session, send_from_directory
from werkzeug.security import generate_password_hash, check_password_hash
import database as db
from database import MESES

app = Flask(__name__)
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'teto-mac-ses-sp-2024-cgof')

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('usuario_id'):
            flash('Faça login para continuar.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('usuario_id'):
            return redirect(url_for('login'))
        if session.get('usuario_perfil') != 'admin':
            flash('Acesso restrito a administradores.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

BASE_TETOS = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'TETOS'))

def _anos_disponiveis():
    """Retorna anos do banco + intervalo 2020..ano_atual+2, ordenado desc."""
    try:
        anos_meses = db.obter_anos_meses()
        anos_db = set(am['ano'] for am in anos_meses)
    except Exception:
        anos_db = set()
    ano_atual = datetime.now().year
    anos_range = set(range(2020, ano_atual + 3))
    return sorted(anos_db | anos_range, reverse=True)

def _competencia_padrao():
    """Ano/mês pré-selecionados na tela de 'Inserir Novo Registro'.

    Calculado a partir do mês corrente do servidor + um deslocamento (em meses)
    configurável pelo admin em /admin/campos (chave 'competencia_offset_meses').
    Padrão: -1 (mês anterior)."""
    try:
        offset = int(db.obter_config('competencia_offset_meses', -1))
    except (TypeError, ValueError):
        offset = -1
    hoje = datetime.now()
    total_meses = (hoje.year * 12 + (hoje.month - 1)) + offset
    ano = total_meses // 12
    mes = total_meses % 12 + 1
    return ano, mes

# ── Favicon ───────────────────────────────────────────────────────────────────

@app.route('/favicon.ico')
def favicon():
    return send_from_directory(
        os.path.join(app.root_path, 'static', 'img'),
        'brasao.png', mimetype='image/png'
    )

# ── Inicialização ─────────────────────────────────────────────────────────────

@app.before_request
def setup():
    db.init_db()

# ── Autenticação ──────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    if session.get('usuario_id'):
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        senha = request.form.get('senha', '')
        try:
            usuario = db.buscar_usuario_por_email(email)
        except Exception as e:
            flash('Tabela de usuários não encontrada. Execute o SQL em criar_tabela_usuarios.sql no Supabase Dashboard e rode setup_admin.py.', 'danger')
            return render_template('login.html', email=email)
        if usuario and check_password_hash(usuario['senha_hash'], senha):
            session['usuario_id'] = usuario['id']
            session['usuario_nome'] = usuario['nome']
            session['usuario_email'] = usuario['email']
            session['usuario_perfil'] = usuario['perfil']
            db.registrar_acesso(usuario['id'])
            return redirect(url_for('dashboard'))
        flash('Email ou senha incorretos.', 'danger')
        return render_template('login.html', email=email)
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ── Alterar Senha (usuário logado) ────────────────────────────────────────────

@app.route('/alterar-senha', methods=['GET', 'POST'])
@login_required
def alterar_senha():
    if request.method == 'POST':
        senha_atual = request.form.get('senha_atual', '')
        nova = request.form.get('nova_senha', '')
        confirmar = request.form.get('confirmar_senha', '')
        usuario = db.buscar_usuario_por_id(session['usuario_id'])
        if not check_password_hash(usuario['senha_hash'], senha_atual):
            flash('Senha atual incorreta.', 'danger')
        elif nova != confirmar:
            flash('As senhas não conferem.', 'danger')
        elif len(nova) < 6:
            flash('A nova senha deve ter pelo menos 6 caracteres.', 'warning')
        else:
            db.atualizar_senha(session['usuario_id'], generate_password_hash(nova))
            flash('Senha alterada com sucesso!', 'success')
            return redirect(url_for('dashboard'))
    return render_template('alterar_senha.html')

# ── Gerenciamento de Usuários (admin) ─────────────────────────────────────────

@app.route('/usuarios')
@admin_required
def usuarios():
    lista = db.listar_usuarios()
    return render_template('usuarios.html', usuarios=lista)

@app.route('/usuarios/criar', methods=['POST'])
@admin_required
def criar_usuario():
    nome = request.form.get('nome', '').strip()
    email = request.form.get('email', '').strip()
    senha = request.form.get('senha', '')
    perfil = request.form.get('perfil', 'usuario')
    if not nome or not email or not senha:
        flash('Preencha todos os campos.', 'danger')
    else:
        try:
            db.criar_usuario(nome, email, generate_password_hash(senha), perfil)
            flash(f'Usuário {nome} criado com sucesso!', 'success')
        except Exception as e:
            flash(f'Erro ao criar usuário: {e}', 'danger')
    return redirect(url_for('usuarios'))

@app.route('/usuarios/editar', methods=['POST'])
@admin_required
def editar_usuario():
    uid = request.form.get('id', type=int)
    nome = request.form.get('nome', '').strip()
    email = request.form.get('email', '').strip()
    perfil = request.form.get('perfil', 'usuario')
    ativo = request.form.get('ativo', '1') == '1'
    try:
        db.editar_usuario_db(uid, nome, email, perfil, ativo)
        flash('Usuário atualizado com sucesso!', 'success')
    except Exception as e:
        flash(f'Erro ao atualizar: {e}', 'danger')
    return redirect(url_for('usuarios'))

@app.route('/usuarios/deletar/<int:id>', methods=['POST'])
@admin_required
def deletar_usuario(id):
    if id == session.get('usuario_id'):
        flash('Você não pode excluir seu próprio usuário.', 'danger')
    else:
        try:
            db.deletar_usuario_db(id)
            flash('Usuário removido.', 'success')
        except Exception as e:
            flash(f'Erro ao remover: {e}', 'danger')
    return redirect(url_for('usuarios'))

@app.route('/usuarios/trocar-senha', methods=['POST'])
@admin_required
def admin_trocar_senha():
    uid = request.form.get('id', type=int)
    nova = request.form.get('nova_senha', '')
    confirmar = request.form.get('confirmar_senha', '')
    if nova != confirmar:
        flash('As senhas não conferem.', 'danger')
    elif len(nova) < 6:
        flash('A senha deve ter pelo menos 6 caracteres.', 'warning')
    else:
        db.atualizar_senha(uid, generate_password_hash(nova))
        flash('Senha alterada com sucesso!', 'success')
    return redirect(url_for('usuarios'))

# ── Dashboard ─────────────────────────────────────────────────────────────────

@app.route('/')
@login_required
def dashboard():
    anos_meses = db.obter_anos_meses()
    anos_disponiveis = sorted(set(am['ano'] for am in anos_meses), reverse=True)

    ano_raw = request.args.get('ano', '')
    todos_anos = (ano_raw == '0' or ano_raw == '')
    ano_sel = 0 if todos_anos else int(ano_raw) if ano_raw.isdigit() else 0
    mes_sel = request.args.get('mes', type=int)

    if not todos_anos and not ano_sel and anos_meses:
        todos_anos = False
        ano_sel = anos_meses[0]['ano']
        mes_sel = anos_meses[0]['mes']
    elif not todos_anos and ano_sel and not mes_sel and anos_meses:
        meses_do_ano = [am['mes'] for am in anos_meses if am['ano'] == ano_sel]
        mes_sel = max(meses_do_ano) if meses_do_ano else 1

    stats = db.estatisticas_gerais()
    evolucao = db.grafico_evolucao_mensal()

    if todos_anos:
        kpis = db.dashboard_kpis_geral()
        kpi_anterior = {}
        por_drs, por_tipo, top_unidades = [], [], []
    else:
        kpis = db.dashboard_kpis(ano_sel, mes_sel)
        por_drs   = db.grafico_por_drs(ano_sel, mes_sel)
        por_tipo  = db.grafico_por_tipo(ano_sel, mes_sel)
        top_unidades = db.grafico_top_unidades(ano_sel, mes_sel)
        kpi_anterior = {}
        if anos_meses and len(anos_meses) > 1:
            am2 = anos_meses[1]
            kpi_anterior = db.dashboard_kpis(am2['ano'], am2['mes'])

    return render_template('dashboard.html',
        kpis=kpis,
        kpi_anterior=kpi_anterior,
        stats=stats,
        todos_anos=todos_anos,
        anos_meses=anos_meses,
        ano_sel=ano_sel,
        mes_sel=mes_sel,
        meses=MESES,
        anos_disponiveis=anos_disponiveis,
        evolucao_json=json.dumps([{
            'ano': r['ano'], 'mes': r['mes'],
            'total': r['total'] or 0,
            'unidades': r['unidades']
        } for r in evolucao]),
        drs_json=json.dumps([{
            'drs': r['drs'], 'total': r['total'] or 0, 'unidades': r['unidades']
        } for r in por_drs]),
        tipo_json=json.dumps([{
            'tipo': r['tipo'], 'total': r['total'] or 0, 'unidades': r['unidades']
        } for r in por_tipo]),
        top_json=json.dumps([{
            'unidade': (r['unidade'] or '')[:40], 'total': r['total'] or 0
        } for r in top_unidades])
    )

# ── Pesquisa ──────────────────────────────────────────────────────────────────

@app.route('/pesquisa')
@login_required
def pesquisa():
    filtros = {k: v for k, v in request.args.items() if v}
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)

    registros, total = db.pesquisar(filtros, page, per_page)
    total_pages = (total + per_page - 1) // per_page

    anos_meses = db.obter_anos_meses()
    drs_lista = db.obter_drs_lista()
    anos_disponiveis = sorted(set(am['ano'] for am in anos_meses), reverse=True)

    return render_template('pesquisa.html',
        registros=registros,
        total=total,
        page=page,
        per_page=per_page,
        total_pages=total_pages,
        filtros=filtros,
        meses=MESES,
        drs_lista=drs_lista,
        anos_disponiveis=anos_disponiveis,
    )

@app.route('/api/autocomplete')
@login_required
def autocomplete():
    campo = request.args.get('campo', '').strip()
    q     = request.args.get('q', '').strip()
    ano   = request.args.get('ano', type=int)
    mes   = request.args.get('mes', type=int)
    if not q or len(q) < 2:
        return jsonify([])
    if campo:
        # nova API: retorna lista de strings para um campo específico
        return jsonify(db.autocomplete_valores(campo, q, ano, mes))
    # legada: retorna lista de dicts {cnes, unidade, municipio} para o form.html
    return jsonify(db.buscar_unidades_autocomplete(q))

@app.route('/api/registro/<int:id>')
@login_required
def api_registro(id):
    registro = db.buscar_registro(id)
    if not registro:
        return jsonify({'erro': 'Registro não encontrado'}), 404
    return jsonify(registro)

# ── CRUD ──────────────────────────────────────────────────────────────────────

@app.route('/inserir', methods=['GET', 'POST'])
@login_required
def inserir():
    if request.method == 'POST':
        dados = _form_para_dict(request.form)
        try:
            new_id = db.inserir_registro(dados)
            flash(f'Registro inserido com sucesso! ID: {new_id}', 'success')
            return redirect(url_for('detalhe', id=new_id))
        except Exception as e:
            flash(f'Erro ao inserir: {e}', 'danger')

    ano_default, mes_default = _competencia_padrao()
    return render_template('form.html',
        registro=None,
        meses=MESES,
        anos_disponiveis=_anos_disponiveis(),
        ano_default=ano_default,
        mes_default=mes_default,
        titulo='Inserir Novo Registro',
        secoes=db.listar_secoes_config(),
        campos=db.listar_campos_config(),
    )

@app.route('/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar(id):
    registro = db.buscar_registro(id)
    if not registro:
        flash('Registro não encontrado', 'danger')
        return redirect(url_for('pesquisa'))

    if request.method == 'POST':
        dados = _form_para_dict(request.form)
        try:
            db.atualizar_registro(id, dados)
            flash('Registro atualizado com sucesso!', 'success')
            return redirect(url_for('detalhe', id=id))
        except Exception as e:
            flash(f'Erro ao atualizar: {e}', 'danger')

    return render_template('form.html',
        registro=registro,
        meses=MESES,
        anos_disponiveis=_anos_disponiveis(),
        titulo=f'Editar Registro #{id}',
        secoes=db.listar_secoes_config(),
        campos=db.listar_campos_config(),
    )

@app.route('/detalhe/<int:id>')
@login_required
def detalhe(id):
    registro = db.buscar_registro(id)
    if not registro:
        flash('Registro não encontrado', 'danger')
        return redirect(url_for('pesquisa'))

    historico = []
    if registro.get('cnes'):
        historico = db.comparativo_unidade(registro['cnes'])

    return render_template('detalhe.html',
        registro=registro,
        meses=MESES,
        historico=historico,
        historico_json=json.dumps([{
            'label': f"{MESES.get(r['mes'],'')} {r['ano']}",
            'total': r['total'] or 0
        } for r in historico])
    )

@app.route('/deletar/<int:id>', methods=['POST'])
@login_required
def deletar(id):
    try:
        db.deletar_registro(id)
        flash('Registro deletado com sucesso!', 'success')
    except Exception as e:
        flash(f'Erro ao deletar: {e}', 'danger')
    return redirect(url_for('pesquisa'))

def _form_para_dict(form):
    # Carrega tipos dos campos para conversão correta
    try:
        campos_cfg = db.listar_campos_config(incluir_inativos=True)
        meta_por_key = {c['campo_key']: c for c in campos_cfg}
        # Lookup também pela coluna_db (ex: 'teto_mac' → campo 'teto_mac_campo')
        meta_por_coluna = {c['coluna_db']: c for c in campos_cfg if c.get('coluna_db')}
    except Exception:
        meta_por_key = {}
        meta_por_coluna = {}

    campos_int = {'ano', 'mes'}
    campos_num_fixos = {'drs'}

    dados = {}
    extras = {}

    for k, v in form.items():
        if k == 'csrf_token':
            continue
        # Busca por campo_key primeiro, depois por coluna_db (para campos com alias)
        meta = meta_por_key.get(k) or meta_por_coluna.get(k)
        if meta:
            tipo = meta.get('tipo', 'moeda')
            coluna_db = meta.get('coluna_db')
            if tipo == 'numero':
                try:
                    val = int(round(float(str(v).replace(',', '.')))) if v else 0
                except Exception:
                    val = 0
            elif tipo in ('moeda', 'calculado'):
                try:
                    val = float(str(v).replace(',', '.')) if v else 0.0
                except Exception:
                    val = 0.0
            else:
                val = v.upper().strip() if tipo == 'texto' else (v.strip() if v else '')
            if coluna_db:
                # Salva sempre pelo nome da coluna real no DB
                dados[coluna_db] = val
            else:
                # Campo personalizado → vai para campos_extras
                extras[k] = val
        elif k in campos_int:
            dados[k] = int(v) if v else 0
        elif k in campos_num_fixos:
            try:
                dados[k] = float(str(v).replace(',', '.')) if v else 0.0
            except Exception:
                dados[k] = 0.0
        else:
            dados[k] = v.upper().strip() if v else ''

    if extras:
        dados['campos_extras'] = extras
    return dados

# ── Relatórios ────────────────────────────────────────────────────────────────

@app.route('/relatorios')
@login_required
def relatorios():
    anos_meses = db.obter_anos_meses()
    anos_disponiveis = sorted(set(am['ano'] for am in anos_meses), reverse=True)
    ultimo = anos_meses[0] if anos_meses else {'ano': 2026, 'mes': 1}
    return render_template('relatorios.html',
        meses=MESES,
        anos_disponiveis=anos_disponiveis,
        ultimo=ultimo
    )

@app.route('/relatorio/resumo-drs')
@login_required
def relatorio_resumo_drs():
    ano = request.args.get('ano', type=int)
    mes = request.args.get('mes', type=int)
    if not ano or not mes:
        anos_meses = db.obter_anos_meses()
        if anos_meses:
            ano, mes = anos_meses[0]['ano'], anos_meses[0]['mes']

    dados = db.relatorio_resumo_drs(ano, mes)
    total_geral = sum(r['total_geral'] or 0 for r in dados)
    anos_meses = db.obter_anos_meses()
    anos_disponiveis = sorted(set(am['ano'] for am in anos_meses), reverse=True)

    return render_template('relatorio_drs.html',
        dados=dados, ano=ano, mes=mes,
        total_geral=total_geral,
        meses=MESES,
        anos_disponiveis=anos_disponiveis,
        anos_meses=anos_meses
    )

@app.route('/relatorio/periodo')
@login_required
def relatorio_periodo():
    ano_ini = request.args.get('ano_ini', type=int, default=2026)
    mes_ini = request.args.get('mes_ini', type=int, default=1)
    ano_fim = request.args.get('ano_fim', type=int, default=2026)
    mes_fim = request.args.get('mes_fim', type=int, default=12)

    dados = db.relatorio_periodo(ano_ini, mes_ini, ano_fim, mes_fim)
    anos_meses = db.obter_anos_meses()
    anos_disponiveis = sorted(set(am['ano'] for am in anos_meses), reverse=True)

    return render_template('relatorio_periodo.html',
        dados=dados,
        ano_ini=ano_ini, mes_ini=mes_ini,
        ano_fim=ano_fim, mes_fim=mes_fim,
        meses=MESES,
        anos_disponiveis=anos_disponiveis
    )

@app.route('/relatorio/comparativo-unidade')
@login_required
def relatorio_comparativo_unidade():
    cnes = request.args.get('cnes', '')
    dados = []
    unidade_nome = ''
    if cnes:
        dados = db.comparativo_unidade(cnes)
        if dados:
            unidade_nome = dados[0].get('unidade', '')

    return render_template('relatorio_unidade.html',
        dados=dados,
        cnes=cnes,
        unidade_nome=unidade_nome,
        meses=MESES,
        historico_json=json.dumps([{
            'label': f"{MESES.get(r['mes'],'')} {r['ano']}",
            'total': r['total'] or 0,
            'aih': (r.get('aih_mc') or 0) + (r.get('aih_ac') or 0),
            'sia': (r.get('sia_mc') or 0) + (r.get('sia_ac') or 0),
        } for r in dados])
    )

# ── Exportação Excel ──────────────────────────────────────────────────────────

@app.route('/exportar/excel')
@login_required
def exportar_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash('openpyxl não instalado', 'danger')
        return redirect(url_for('pesquisa'))

    filtros = {k: v for k, v in request.args.items() if v and k not in ('format',)}
    registros, total = db.pesquisar(filtros, page=1, per_page=99999)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Teto MAC'

    azul_header = PatternFill("solid", fgColor="1e3a5f")
    fonte_header = Font(bold=True, color="FFFFFF", size=11)
    fonte_titulo = Font(bold=True, color="1e3a5f", size=14)
    borda = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.merge_cells('A1:AK1')
    ws['A1'] = 'SECRETARIA DE ESTADO DA SAÚDE - SP | CONTROLE DE TETO MAC'
    ws['A1'].font = fonte_titulo
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25

    headers = [
        'Ano', 'Mês', 'DRS', 'Tipo', 'HU', 'Município', 'CNES', 'CNPJ', 'Unidade',
        'AIH Físico', 'AIH FAEC', 'SIA FAEC', 'Equip. Hemodiálise',
        'AIH MC', 'AIH AC', 'AIH Total', 'SIA MC', 'SIA AC', 'SIA Total',
        'Teto Global', 'Teto MC', 'Teto AC', 'Teto MAC', 'Total Teto MAC',
        'IntegraSUS', 'IAC', '100% SUS', 'OPO', 'Rede Viver',
        'RSME', 'RCE/RCEG', 'RAU/SOS', 'RCA/RCAN', 'IAPI',
        'Resid. Médica', 'Melhor em Casa', 'CER', 'Doenças Raras',
        'Of. Ortopédica', 'IHAC', 'Total MC+AC+Incentivos'
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = fonte_header
        cell.fill = azul_header
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = borda
    ws.row_dimensions[2].height = 30

    campos = [
        'ano', 'mes', 'drs', 'tipo', 'hu', 'municipio', 'cnes', 'cnpj', 'unidade',
        'aih_fisico', 'aih_faec', 'sia_faec', 'equip_hemodialise',
        'aih_mc', 'aih_ac', 'aih_total', 'sia_mc', 'sia_ac', 'sia_total',
        'teto_global', 'teto_mc', 'teto_ac', 'teto_mac', 'total_teto_mac',
        'integrasus', 'iac', 'sus_100', 'opo', 'rede_viver_sem_limite',
        'rsme', 'rce_rceg', 'rau_hosp_sos', 'rca_rcan', 'iapi',
        'residencia_medica', 'melhor_em_casa', 'cer', 'doencas_raras',
        'oficina_ortopedica', 'ihac', 'total_mc_ac_incentivos'
    ]

    fill_par = PatternFill("solid", fgColor="EBF3FB")
    fmt_moeda = '#,##0.00'
    fmt_numero = '#,##0'
    col_aih_fisico = campos.index('aih_fisico') + 1

    for row_num, reg in enumerate(registros, 3):
        fill = fill_par if row_num % 2 == 0 else None
        for col, campo in enumerate(campos, 1):
            val = reg.get(campo, '')
            if campo == 'mes' and val:
                val = MESES.get(int(val), val)
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border = borda
            if fill:
                cell.fill = fill
            if col > 9 and isinstance(val, (int, float)) and val:
                cell.number_format = fmt_numero if col == col_aih_fisico else fmt_moeda
                cell.alignment = Alignment(horizontal='right')

    # Larguras
    larguras = [6, 10, 5, 30, 5, 20, 10, 16, 40] + [14] * 32
    for col, w in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = 'A3'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    ano_f = filtros.get('ano', 'todos')
    mes_f = MESES.get(int(filtros.get('mes', 0)), 'todos') if filtros.get('mes') else 'todos'
    filename = f'TetMAC_{ano_f}_{mes_f}.xlsx'

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/exportar/excel-drs')
@login_required
def exportar_excel_drs():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ano = request.args.get('ano', type=int)
    mes = request.args.get('mes', type=int)
    dados = db.relatorio_resumo_drs(ano, mes)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Resumo DRS {MESES.get(mes,"")} {ano}'

    azul = PatternFill("solid", fgColor="1e3a5f")
    fonte_h = Font(bold=True, color="FFFFFF")

    ws['A1'] = f'RESUMO POR DRS - {MESES.get(mes,"").upper()} {ano}'
    ws['A1'].font = Font(bold=True, size=13, color="1e3a5f")
    ws.merge_cells('A1:I1')

    headers = ['DRS', 'Total Unidades', 'AIH Físico', 'Total AIH', 'Total SIA', 'Teto MAC', 'Total Incentivos', 'Total Geral']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = fonte_h
        c.fill = azul

    for r, row in enumerate(dados, 3):
        ws.cell(r, 1, row.get('drs'))
        ws.cell(r, 2, row.get('total_unidades'))
        ws.cell(r, 3, row.get('aih_fisico'))
        ws.cell(r, 4, row.get('total_aih'))
        ws.cell(r, 5, row.get('total_sia'))
        ws.cell(r, 6, row.get('teto_mac'))
        ws.cell(r, 7, row.get('total_incentivos'))
        ws.cell(r, 8, row.get('total_geral'))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'ResumoD RS_{ano}_{mes:02d}.xlsx'
    )

# ── Importação ────────────────────────────────────────────────────────────────

@app.route('/importar', methods=['GET', 'POST'])
@login_required
def importar():
    from import_xls import importar_arquivo_xls, importar_todos_finais

    if request.method == 'POST':
        acao = request.form.get('acao')
        substituir = request.form.get('substituir') == '1'

        if acao == 'importar_multiplos':
            arquivos = request.files.getlist('arquivos')
            if not arquivos or all(not a.filename for a in arquivos):
                flash('Nenhum arquivo selecionado.', 'danger')
            else:
                resultados = []
                for arquivo in arquivos:
                    if not arquivo.filename:
                        continue
                    import tempfile
                    ext = os.path.splitext(arquivo.filename)[1]
                    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
                    arquivo.save(tmp.name)
                    tmp.close()
                    try:
                        res = importar_arquivo_xls(tmp.name, None, None, substituir, nome_original=arquivo.filename)
                        res['arquivo'] = arquivo.filename
                        resultados.append(res)
                    finally:
                        os.unlink(tmp.name)
                total_imp = sum(r['importados'] for r in resultados)
                total_err = sum(r['erros'] for r in resultados)
                flash(f'{len(resultados)} arquivo(s) processado(s) — {total_imp} registros importados, {total_err} erros.',
                      'success' if total_err == 0 else 'warning')
                return render_template('importar.html',
                    resultados=resultados,
                    anos_disponiveis=list(range(2021, 2028)),
                    meses=MESES,
                    historico_imp=_obter_historico_importacoes()
                )

        elif acao == 'importar_arquivo':
            arquivo = request.files.get('arquivo')
            if not arquivo:
                flash('Nenhum arquivo selecionado', 'danger')
            else:
                import tempfile
                tmp = tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(arquivo.filename)[1])
                arquivo.save(tmp.name)
                tmp.close()
                try:
                    ano = request.form.get('ano', type=int)
                    mes = request.form.get('mes', type=int)
                    res = importar_arquivo_xls(tmp.name, ano, mes, substituir, nome_original=arquivo.filename)
                    res['arquivo'] = arquivo.filename
                    flash(f'Arquivo "{arquivo.filename}": {res["importados"]} registros importados de {res["total"]} ({res["pulados"]} pulados, {res["erros"]} erros)', 'success' if res['erros'] == 0 else 'warning')
                    return render_template('importar.html',
                        resultados=[res],
                        anos_disponiveis=list(range(2022, 2027)),
                        meses=MESES
                    )
                finally:
                    os.unlink(tmp.name)

    historico_imp = _obter_historico_importacoes()
    return render_template('importar.html',
        resultados=None,
        anos_disponiveis=list(range(2022, 2027)),
        meses=MESES,
        historico_imp=historico_imp
    )

def _obter_historico_importacoes():
    from config import USE_SUPABASE
    if USE_SUPABASE:
        from database import get_sb
        r = get_sb().table('importacoes').select('*').order('created_at', desc=True).limit(50).execute()
        return r.data if r.data else []
    else:
        conn = db.get_db()
        rows = conn.execute(
            "SELECT * FROM importacoes ORDER BY created_at DESC LIMIT 50"
        ).fetchall()
        conn.close()
        return [dict(r) for r in rows]

# ── Auditoria ─────────────────────────────────────────────────────────────────

@app.route('/auditoria')
@admin_required
def auditoria():
    anos_meses = db.obter_anos_meses()
    anos_disponiveis = sorted(set(am['ano'] for am in anos_meses), reverse=True)
    ano = request.args.get('ano', type=int)
    mes = request.args.get('mes', type=int)
    if not ano and anos_meses:
        ano, mes = anos_meses[0]['ano'], anos_meses[0]['mes']
    drs_lista = db.obter_drs_lista()
    return render_template('auditoria.html',
        anos_meses=anos_meses, anos_disponiveis=anos_disponiveis,
        ano_sel=ano, mes_sel=mes, meses=MESES, drs_lista=drs_lista
    )

@app.route('/api/auditoria/validacao')
@admin_required
def api_auditoria_validacao():
    ano = request.args.get('ano', type=int)
    mes = request.args.get('mes', type=int)
    if not ano or not mes:
        return jsonify({'error': 'ano e mes obrigatorios'}), 400
    return jsonify(db.auditoria_validacao(ano, mes))

@app.route('/api/auditoria/registros')
@admin_required
def api_auditoria_registros():
    ano = request.args.get('ano', type=int)
    mes = request.args.get('mes', type=int)
    drs = request.args.get('drs', type=int)
    busca = request.args.get('q', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = min(request.args.get('per_page', 100, type=int), 500)
    if not ano or not mes:
        return jsonify({'registros': [], 'total': 0})
    regs, total = db.auditoria_registros(ano, mes, drs or None, busca or None, page, per_page)
    return jsonify({'registros': regs, 'total': total, 'page': page, 'per_page': per_page})

@app.route('/auditoria/deletar-registros', methods=['POST'])
@admin_required
def auditoria_deletar_registros():
    data = request.get_json()
    ids  = data.get('ids', [])
    if not ids:
        return jsonify({'ok': False, 'msg': 'Nenhum ID fornecido'}), 400
    try:
        n = db.auditoria_deletar_ids(ids)
        return jsonify({'ok': True, 'deletados': n})
    except Exception as e:
        return jsonify({'ok': False, 'msg': str(e)}), 500

@app.route('/auditoria/deletar-periodo', methods=['POST'])
@admin_required
def auditoria_deletar_periodo():
    data = request.get_json()
    ano, mes = data.get('ano'), data.get('mes')
    if not ano or not mes:
        return jsonify({'ok': False, 'msg': 'ano e mes obrigatorios'}), 400
    try:
        n = db.auditoria_deletar_periodo(ano, mes)
        return jsonify({'ok': True, 'deletados': n})
    except Exception as e:
        return jsonify({'ok': False, 'msg': str(e)}), 500

@app.route('/auditoria/comparar', methods=['POST'])
@admin_required
def auditoria_comparar():
    from import_xls import extrair_ano_mes_do_nome, mapear_colunas, val_num, val_str
    arquivo = request.files.get('arquivo')
    if not arquivo:
        return jsonify({'error': 'Arquivo não enviado'}), 400
    ano = request.form.get('ano', type=int)
    mes = request.form.get('mes', type=int)
    import tempfile
    ext = os.path.splitext(arquivo.filename)[1]
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
    arquivo.save(tmp.name)
    tmp.close()
    try:
        if not ano or not mes:
            ano, mes = extrair_ano_mes_do_nome(arquivo.filename)
        registros_xls = _ler_planilha_auditoria(tmp.name, ano, mes)
        resultado = db.auditoria_comparar(registros_xls, ano, mes)
        resultado.update({'ano': ano, 'mes': mes, 'arquivo': arquivo.filename})
        return jsonify(resultado)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        os.unlink(tmp.name)

def _ler_planilha_auditoria(filepath, ano, mes):
    """Lê planilha e retorna lista de dicts sem salvar no banco."""
    from import_xls import mapear_colunas, val_num, val_str
    import xlrd, openpyxl
    registros = []
    try:
        try:
            wb = xlrd.open_workbook(filepath)
            ws = wb.sheet_by_index(0)
            header_row = None
            for r in range(min(5, ws.nrows)):
                if 'DRS' in [str(ws.cell_value(r, c)).strip().upper() for c in range(ws.ncols)]:
                    header_row = r; break
            if header_row is None:
                return []
            mapa = mapear_colunas([ws.cell_value(header_row, c) for c in range(ws.ncols)])
            for r in range(header_row + 1, ws.nrows):
                rv = [ws.cell_value(r, c) for c in range(ws.ncols)]
                if not str(rv[mapa.get('drs', 0)] if 'drs' in mapa else '').strip() and \
                   not str(rv[mapa.get('unidade', 0)] if 'unidade' in mapa else '').strip():
                    continue
                g = lambda f: rv[mapa[f]] if f in mapa and mapa[f] < len(rv) else 0
                cnes = val_str(rv[mapa['cnes']]) if 'cnes' in mapa else ''
                if cnes and cnes != 'RESREC':
                    try: cnes = str(int(float(cnes)))
                    except: pass
                registros.append({
                    'cnes': cnes,
                    'unidade': val_str(g('unidade')).upper(),
                    'municipio': val_str(g('municipio')).upper(),
                    'drs': val_num(g('drs')),
                    'total_mc_ac_incentivos': val_num(g('total_mc_ac_incentivos')),
                    'teto_mac': val_num(g('teto_mac')),
                    'aih_mc': val_num(g('aih_mc')), 'aih_ac': val_num(g('aih_ac')),
                    'sia_mc': val_num(g('sia_mc')), 'sia_ac': val_num(g('sia_ac')),
                })
        except Exception:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            header_row = next((i for i, r in enumerate(rows[:5])
                               if 'DRS' in [str(v).strip().upper() for v in r if v]), None)
            if header_row is None:
                return []
            mapa = mapear_colunas(list(rows[header_row]))
            for rv in rows[header_row + 1:]:
                if not rv or all(v is None for v in rv[:5]):
                    continue
                g = lambda f: rv[mapa[f]] if f in mapa and mapa[f] < len(rv) else 0
                cnes = val_str(rv[mapa['cnes']]) if 'cnes' in mapa else ''
                registros.append({
                    'cnes': cnes,
                    'unidade': val_str(g('unidade')).upper(),
                    'total_mc_ac_incentivos': val_num(g('total_mc_ac_incentivos')),
                    'teto_mac': val_num(g('teto_mac')),
                    'aih_mc': val_num(g('aih_mc')), 'aih_ac': val_num(g('aih_ac')),
                    'sia_mc': val_num(g('sia_mc')), 'sia_ac': val_num(g('sia_ac')),
                })
    except Exception as e:
        pass
    return registros

# ── Gráficos (API JSON) ────────────────────────────────────────────────────────

@app.route('/api/graficos/evolucao')
def api_evolucao():
    anos = request.args.getlist('anos', type=int)
    dados = db.grafico_evolucao_mensal(anos if anos else None)
    return jsonify(dados)

@app.route('/api/graficos/drs')
def api_drs():
    ano = request.args.get('ano', type=int)
    mes = request.args.get('mes', type=int)
    dados = db.grafico_por_drs(ano, mes)
    return jsonify(dados)

@app.route('/api/graficos/tipo')
def api_tipo():
    ano = request.args.get('ano', type=int)
    mes = request.args.get('mes', type=int)
    dados = db.grafico_por_tipo(ano, mes)
    return jsonify(dados)

@app.route('/api/graficos/top-unidades')
def api_top_unidades():
    ano = request.args.get('ano', type=int)
    mes = request.args.get('mes', type=int)
    limite = request.args.get('limite', 15, type=int)
    dados = db.grafico_top_unidades(ano, mes, limite)
    return jsonify(dados)

@app.route('/api/kpis')
def api_kpis():
    ano = request.args.get('ano', type=int)
    mes = request.args.get('mes', type=int)
    return jsonify(db.dashboard_kpis(ano, mes))

@app.route('/central-analitica')
@login_required
def central_analitica():
    anos_meses = db.obter_anos_meses()
    anos_disponiveis = sorted(set(am['ano'] for am in anos_meses), reverse=True)
    ultimo = anos_meses[0] if anos_meses else {'ano': 2026, 'mes': 1}
    return render_template('central_analitica.html',
        meses=MESES,
        anos_disponiveis=anos_disponiveis,
        ultimo=ultimo
    )

@app.route('/api/kpis-central')
@login_required
def api_kpis_central():
    ano = request.args.get('ano', type=int)
    mes = request.args.get('mes', type=int)
    return jsonify(db.kpis_central(ano, mes))

@app.route('/api/analitico', methods=['GET', 'POST'])
@login_required
def api_analitico():
    if request.method == 'POST':
        body = request.get_json(force=True) or {}
    else:
        body = request.args.to_dict()
    ano  = body.get('ano') or request.args.get('ano', type=int)
    mes  = body.get('mes') or request.args.get('mes', type=int)
    try:
        ano = int(ano); mes = int(mes)
    except (TypeError, ValueError):
        return jsonify({'error': 'ano e mes obrigatórios'}), 400
    dimensoes   = body.get('dimensoes', [])
    metricas    = body.get('metricas', ['total_mc_ac_incentivos'])
    filtros     = body.get('filtros', {})
    ordenar_por = body.get('ordenar_por')
    limite      = int(body.get('limite', 500))
    return jsonify(db.consulta_analitica(ano, mes, dimensoes, metricas, filtros, ordenar_por, limite))

@app.route('/api/relatorio/unidade')
@login_required
def api_relatorio_unidade():
    ano = request.args.get('ano', type=int)
    mes = request.args.get('mes', type=int)
    return jsonify(db.relatorio_por_unidade(ano, mes))

@app.route('/api/relatorio/municipio')
@login_required
def api_relatorio_municipio():
    ano = request.args.get('ano', type=int)
    mes = request.args.get('mes', type=int)
    return jsonify(db.relatorio_por_municipio(ano, mes))

@app.route('/api/relatorio/fundo')
@login_required
def api_relatorio_fundo():
    ano = request.args.get('ano', type=int)
    mes = request.args.get('mes', type=int)
    return jsonify(db.relatorio_fundo(ano, mes))

@app.route('/api/relatorio/incentivos')
@login_required
def api_relatorio_incentivos():
    ano = request.args.get('ano', type=int)
    mes = request.args.get('mes', type=int)
    return jsonify(db.relatorio_incentivos(ano, mes))

# ── Gráficos dedicados ────────────────────────────────────────────────────────

@app.route('/graficos')
@login_required
def graficos():
    anos_meses = db.obter_anos_meses()
    ano_sel = request.args.get('ano', type=int)
    mes_sel = request.args.get('mes', type=int)
    if not ano_sel and anos_meses:
        ano_sel = anos_meses[0]['ano']
        mes_sel = anos_meses[0]['mes']
    anos_disponiveis = sorted(set(am['ano'] for am in anos_meses), reverse=True)
    return render_template('graficos.html',
        anos_meses=anos_meses,
        ano_sel=ano_sel,
        mes_sel=mes_sel,
        meses=MESES,
        anos_disponiveis=anos_disponiveis
    )

# ── Detalhamento Completo ─────────────────────────────────────────────────────

@app.route('/detalhamento')
@login_required
def detalhamento():
    anos_meses = db.obter_anos_meses()
    anos_disponiveis = sorted(set(am['ano'] for am in anos_meses), reverse=True)
    ano = request.args.get('ano', type=int)
    mes = request.args.get('mes', type=int)
    if not ano and anos_meses:
        ano, mes = anos_meses[0]['ano'], anos_meses[0]['mes']
    drs_lista = db.obter_drs_lista()
    tipos = db.detalhamento_tipos(ano, mes) if ano and mes else []
    return render_template('detalhamento.html',
        anos_meses=anos_meses, anos_disponiveis=anos_disponiveis,
        ano_sel=ano, mes_sel=mes, meses=MESES, drs_lista=drs_lista, tipos=tipos
    )

@app.route('/api/detalhamento/registros')
@login_required
def api_detalhamento_registros():
    ano  = request.args.get('ano', type=int)
    mes  = request.args.get('mes', type=int)
    drs  = request.args.get('drs', type=int)
    tipo = request.args.get('tipo', '').strip() or None
    busca = request.args.get('q', '').strip() or None
    page = request.args.get('page', 1, type=int)
    per_page = min(request.args.get('per_page', 50, type=int), 200)
    sort_col = request.args.get('sort', 'drs').strip()
    sort_dir = request.args.get('sort_dir', 'asc').strip()
    col_filters = {}
    for k, v in request.args.items():
        if not v.strip(): continue
        if k.startswith('cf_'):
            col_filters[k[3:]] = v.strip()
    if not ano or not mes:
        return jsonify({'registros': [], 'total': 0})
    regs, total = db.detalhamento_registros(
        ano, mes, drs or None, tipo, busca, page, per_page, sort_col, sort_dir,
        col_filters or None
    )
    return jsonify({'registros': regs, 'total': total, 'page': page, 'per_page': per_page})

@app.route('/api/detalhamento/valores-unicos')
@login_required
def api_detalhamento_valores_unicos():
    col = request.args.get('col', '').strip()
    ano = request.args.get('ano', type=int)
    mes = request.args.get('mes', type=int)
    if not col or not ano or not mes:
        return jsonify([])
    vals = db.detalhamento_valores_unicos(col, ano, mes)
    return jsonify(vals)

@app.route('/detalhamento/exportar')
@login_required
def detalhamento_exportar():
    import csv
    ano  = request.args.get('ano', type=int)
    mes  = request.args.get('mes', type=int)
    drs  = request.args.get('drs', type=int)
    tipo = request.args.get('tipo', '').strip() or None
    busca = request.args.get('q', '').strip() or None
    if not ano or not mes:
        flash('Selecione um período para exportar.', 'warning')
        return redirect(url_for('detalhamento'))
    regs, _ = db.detalhamento_registros(ano, mes, drs or None, tipo, busca, page=1, per_page=5000)
    output = io.StringIO()
    if regs:
        writer = csv.DictWriter(output, fieldnames=list(regs[0].keys()))
        writer.writeheader()
        writer.writerows(regs)
    output.seek(0)
    nome = f"teto_mac_{ano}_{mes:02d}.csv"
    return Response(output.getvalue(), mimetype='text/csv;charset=utf-8',
                    headers={'Content-Disposition': f'attachment;filename={nome}'})

# ── Portarias ────────────────────────────────────────────────────────────────

def _comprimir_pdf_bytes(input_bytes):
    """Comprime PDF em memória com pikepdf. Retorna bytes comprimidos ou None."""
    try:
        import pikepdf, io
        inp = io.BytesIO(input_bytes)
        out = io.BytesIO()
        with pikepdf.open(inp) as pdf:
            pdf.save(out,
                     compress_streams=True,
                     object_stream_mode=pikepdf.ObjectStreamMode.generate)
        out.seek(0)
        return out.read()
    except Exception:
        return None

@app.route('/api/portarias/<cnes>')
@login_required
def api_portarias(cnes):
    return jsonify(db.listar_portarias(cnes))

@app.route('/portaria/upload', methods=['POST'])
@login_required
def portaria_upload():
    import time
    cnes      = request.form.get('cnes', '').strip()
    descricao = request.form.get('descricao', '').strip()
    arquivo   = request.files.get('arquivo')

    if not cnes:
        return jsonify({'ok': False, 'msg': 'CNES obrigatório'}), 400
    if not arquivo or not arquivo.filename:
        return jsonify({'ok': False, 'msg': 'Arquivo não enviado'}), 400
    if not arquivo.filename.lower().endswith('.pdf'):
        return jsonify({'ok': False, 'msg': 'Apenas arquivos PDF são aceitos'}), 400

    original_bytes      = arquivo.read()
    tamanho_original_kb = max(1, len(original_bytes) // 1024)

    compressed = _comprimir_pdf_bytes(original_bytes)
    if compressed and len(compressed) < len(original_bytes):
        final_bytes = compressed
    else:
        final_bytes = original_bytes

    tamanho_kb   = max(1, len(final_bytes) // 1024)
    storage_path = f"{cnes}/{int(time.time())}.pdf"

    try:
        db.upload_portaria_storage(storage_path, final_bytes)
    except Exception as e:
        return jsonify({'ok': False, 'msg': f'Erro ao salvar no storage: {e}'}), 500

    pid = db.salvar_portaria(
        cnes, arquivo.filename, storage_path,
        tamanho_kb, tamanho_original_kb, descricao
    )
    return jsonify({'ok': True, 'portaria': db.buscar_portaria(pid)})

@app.route('/portaria/<int:pid>/ver')
@login_required
def portaria_ver(pid):
    p = db.buscar_portaria(pid)
    if not p:
        return 'Portaria não encontrada', 404
    try:
        file_bytes = db.download_portaria_storage(p['storage_path'])
    except Exception as e:
        return f'Erro ao baixar arquivo do storage: {e}', 500
    nome = p.get('nome_original', 'portaria.pdf')
    return Response(
        file_bytes,
        mimetype='application/pdf',
        headers={
            'Content-Disposition': f'inline; filename="{nome}"',
            'Content-Length': str(len(file_bytes)),
        }
    )

@app.route('/portaria/<int:pid>/validar', methods=['POST'])
@login_required
def portaria_validar(pid):
    p = db.buscar_portaria(pid)
    if not p:
        return jsonify({'ok': False, 'msg': 'Portaria não encontrada'}), 404
    if p['validado']:
        db.desvalidar_portaria(pid)
        return jsonify({'ok': True, 'validado': False})
    db.validar_portaria(pid, session.get('usuario_nome', 'Sistema'))
    p2 = db.buscar_portaria(pid)
    return jsonify({'ok': True, 'validado': True,
                    'validado_em': p2['validado_em'],
                    'validado_por': p2['validado_por']})

@app.route('/portaria/<int:pid>/deletar', methods=['POST'])
@login_required
def portaria_deletar(pid):
    p = db.deletar_portaria_db(pid)
    if not p:
        return jsonify({'ok': False, 'msg': 'Portaria não encontrada'}), 404
    return jsonify({'ok': True})

# ── Admin: Configuração de Campos ─────────────────────────────────────────────

@app.route('/admin/campos')
@admin_required
def admin_campos():
    secoes = db.listar_secoes_config()
    campos = db.listar_campos_config(incluir_inativos=True)
    campos_por_secao = {}
    for c in campos:
        campos_por_secao.setdefault(c['secao_key'], []).append(c)
    try:
        competencia_offset = int(db.obter_config('competencia_offset_meses', -1))
    except (TypeError, ValueError):
        competencia_offset = -1
    return render_template('admin_campos.html',
        secoes=secoes,
        campos_por_secao=campos_por_secao,
        titulo='Gerenciar Campos do Formulário',
        competencia_offset=competencia_offset,
    )

@app.route('/admin/config/salvar', methods=['POST'])
@admin_required
def admin_config_salvar():
    try:
        offset = int(request.form.get('competencia_offset_meses', -1))
        db.salvar_config('competencia_offset_meses', str(offset))
        flash('Configuração salva com sucesso!', 'success')
    except Exception as e:
        flash(f'Erro ao salvar configuração: {e}', 'danger')
    return redirect(url_for('admin_campos'))

@app.route('/admin/campos/salvar', methods=['POST'])
@admin_required
def admin_campos_salvar():
    dados = {
        'secao_key':   request.form.get('secao_key', '').strip(),
        'campo_key':   request.form.get('campo_key', '').strip().lower().replace(' ', '_'),
        'label':       request.form.get('label', '').strip(),
        'tipo':        request.form.get('tipo', 'moeda'),
        'ordem':       int(request.form.get('ordem', 0) or 0),
        'ativo':       request.form.get('ativo') == '1',
        'obrigatorio': request.form.get('obrigatorio') == '1',
        'formula':     request.form.get('formula', '').strip() or None,
        'coluna_db':   request.form.get('coluna_db', '').strip() or None,
    }
    id_ = request.form.get('id', '').strip()
    if id_:
        dados['id'] = int(id_)
    if not dados['campo_key'] or not dados['label']:
        flash('Campo Key e Label são obrigatórios.', 'danger')
        return redirect(url_for('admin_campos'))
    try:
        db.salvar_campo_config(dados)
        flash('Campo salvo com sucesso!', 'success')
    except Exception as e:
        flash(f'Erro ao salvar campo: {e}', 'danger')
    return redirect(url_for('admin_campos'))

@app.route('/admin/campos/<int:id>/toggle', methods=['POST'])
@admin_required
def admin_campos_toggle(id):
    ativo = request.form.get('ativo') == '1'
    try:
        db.salvar_campo_config({'id': id, 'ativo': ativo})
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'msg': str(e)}), 500

@app.route('/admin/campos/<int:id>/deletar', methods=['POST'])
@admin_required
def admin_campos_deletar(id):
    try:
        db.deletar_campo_config(id)
        flash('Campo removido.', 'success')
    except Exception as e:
        flash(f'Erro: {e}', 'danger')
    return redirect(url_for('admin_campos'))

@app.route('/admin/campos/reordenar', methods=['POST'])
@admin_required
def admin_campos_reordenar():
    try:
        items = request.get_json()
        db.reordenar_campos(items)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'msg': str(e)}), 500

@app.route('/admin/secoes/salvar', methods=['POST'])
@admin_required
def admin_secoes_salvar():
    dados = {
        'secao_key': request.form.get('secao_key', '').strip(),
        'label':     request.form.get('label', '').strip(),
        'cor':       request.form.get('cor', 'primary'),
        'icone':     request.form.get('icone', 'list').strip(),
        'ordem':     int(request.form.get('ordem', 0) or 0),
        'ativo':     request.form.get('ativo') == '1',
    }
    id_ = request.form.get('id', '').strip()
    if id_:
        dados['id'] = int(id_)
    try:
        db.salvar_secao_config(dados)
        flash('Seção salva com sucesso!', 'success')
    except Exception as e:
        flash(f'Erro ao salvar seção: {e}', 'danger')
    return redirect(url_for('admin_campos'))

# ── Filtros Jinja ──────────────────────────────────────────────────────────────

@app.template_filter('moeda')
def moeda_filter(value):
    if value is None:
        return 'R$ 0,00'
    try:
        return f"R$ {float(value):_.2f}".replace('.', ',').replace('_', '.')
    except:
        return 'R$ 0,00'

@app.template_filter('numero')
def numero_filter(value):
    if value is None:
        return '0'
    try:
        return f"{int(round(float(value))):_}".replace('_', '.')
    except Exception:
        return '0'

@app.template_filter('mes_nome')
def mes_nome_filter(value):
    try:
        return MESES.get(int(value), str(value))
    except:
        return str(value)

@app.template_filter('percent_var')
def percent_var_filter(atual, anterior):
    try:
        if not anterior or anterior == 0:
            return 0
        return ((float(atual) - float(anterior)) / float(anterior)) * 100
    except:
        return 0

if __name__ == '__main__':
    db.init_db()
    print("\n" + "="*60)
    print("  SISTEMA TETO MAC - SES-SP")
    print("  Acesse: http://localhost:5000")
    print("="*60 + "\n")
    app.run(debug=True, port=5000, host='0.0.0.0')
